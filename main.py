import random
from classes import InputProcessor
from classes import Player
import copy
import math
import openpyxl


def fitness(teams, information):
    team_size = []
    team_score = []
    for i in range(information.num_teams):
        team_size.append(teams.count(i))
        team_score.append(sum([information.participants[j].value if teams[j] == i else 0 for j in range(len(teams))]))

    if max(team_size) - min(team_size) > information.num_teams:
        return 0

    return (1 / (max(team_score) - min(team_score))) if max(team_score) != min(team_score) else float('inf')


def get_initial_teams(information):
    teams = []
    for participant in range(len(information.participants)):
        teams.append(participant % information.num_teams)
    return teams


def get_neighbours(teams):
    temp_neighbours = []
    for participant1 in range(len(teams)):
        for j in range(participant1 + 1, len(teams)):
            new_neighbour = copy.deepcopy(teams)
            new_neighbour[participant1], new_neighbour[j] = new_neighbour[j], new_neighbour[participant1]
            temp_neighbours.append(new_neighbour)
    return temp_neighbours

def get_teams(teams, information):
    teams_sep = [[] for i in range(information.num_teams)]

    for i in range(len(information.participants)):
        teams_sep[teams[i]].append(information.participants[i])

    for team in teams_sep:
        team.sort()

    return teams_sep


comp_information = InputProcessor('tugofwarparticipants.xlsx', 4)
print(comp_information.participants)

initial_teams = current_teams = get_initial_teams(comp_information)
initial_temperature = current_temperature = 10
cooling_rate = 0.95
num_iterations = 1000

for i in range(num_iterations):
    print(i)
    current_temperature *= cooling_rate
    neighbours = get_neighbours(current_teams)
    fitness_scores = [fitness(neighbour, comp_information) for neighbour in neighbours]
    best_fitness = max(fitness_scores)
    best_teams = neighbours[fitness_scores.index(best_fitness)]
    current_fitness = fitness(current_teams, comp_information)

    if best_fitness >= current_fitness:
        current_teams = best_teams
    else:
        probability = math.exp((best_fitness - current_fitness) / current_temperature)
        if random.random() <= probability:
            current_teams = best_teams

#print(get_teams(current_teams, comp_information), 1/fitness(current_teams, comp_information), [len(team) for team in get_teams(current_teams, comp_information)])

final_teams = get_teams(current_teams, comp_information)

wb = openpyxl.Workbook()
sheet = wb.active
for index1, team in enumerate(final_teams):
    sheet.cell(row = 1, column=index1 * 3 + 1).value = "Team" + str(index1 + 1)
    for index2, participant in enumerate(team):
        sheet.cell(row=index2 + 3, column = index1 * 3 + 1).value = participant.name
        sheet.cell(row = index2 + 3, column= index1 * 3 + 2).value = participant.weight

    sheet.cell(row = max([len(t) for t in final_teams]) + 4, column= index1 * 3 + 2).value = sum([participant.weight for participant in team])

wb.save('tug0fwarteams.xlsx')