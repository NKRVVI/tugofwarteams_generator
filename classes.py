import openpyxl


class InputProcessor:
    def __init__(self, file_path, num_teams):
        self.file = openpyxl.load_workbook(file_path, data_only=True).active
        self.num_teams = num_teams
        self.participants = self.get_participants()

    def get_participants(self):
        participants = []
        for i in range(2, self.file.max_row + 1):
            participants.append(Player(self.file.cell(row=i, column=2).value, self.file.cell(row=i, column=3).value, self.file.cell(row=i, column=4).value))

        weights = [participant.weight for participant in participants]
        min_val = min(weights) - 1
        max_val = max(weights) + 1

        for participant in participants:
            weight = (participant.weight - min_val) * (max_val - min_val) / 5
            participant.value = (weight + participant.strength) / (weight * participant.strength)

        participants.sort(reverse=True)
        return participants


class Player:
    def __init__(self, name, weight, strength):
        self.name = name
        self.weight = weight
        self.strength = strength
        self.value = -1

    def __lt__(self, other):
        return self.value < other.value

    def __gt__(self, other):
        return self.value > other.value

    def __str__(self):
        return f"[{self.name}, {self.weight}]"

    def __repr__(self):
        return self.__str__()
